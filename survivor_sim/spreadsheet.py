"""
Spreadsheet exporter for Survivor simulation.

Usage:
    from survivor_sim.spreadsheet import SpreadsheetRecorder
    rec = SpreadsheetRecorder(sim)
    sim.run()  # or otherwise run simulation
    rec.export_xlsx("season.xlsx")

This recorder listens to sim.events ("tribal_start", "vote_cast", "eliminated")
and builds a per-round record. It can export to .xlsx using openpyxl
(if available) with basic formatting suitable for importing into Google Sheets.
If openpyxl is not installed, it falls back to a CSV-style export (matrix of votes).

The generated spreadsheet contains:
- Column A: player names
- Columns B..: episodes (tribal councils)
- Top area: Eliminated name per episode and vote summary (counts)

"""
from __future__ import annotations

from collections import Counter
from collections.abc import Iterable
from typing import Any, Optional, TypedDict, cast


class RoundRecord(TypedDict):
    round: Optional[int]
    pool: list[str]
    votes: dict[str, str]
    eliminated: Optional[str]


def _round_value(value: object) -> Optional[int]:
    return value if isinstance(value, int) else None


def _name_of(value: object) -> str:
    name = getattr(value, "name", None)
    return name if isinstance(name, str) else str(value)


def _iter_objects(value: object) -> Iterable[object]:
    if value is None or isinstance(value, (str, bytes)):
        return ()
    if isinstance(value, Iterable):
        return cast(Iterable[object], value)
    return ()


class SpreadsheetRecorder:
    def __init__(self, sim: Any) -> None:
        """Attach to a `SurvivorSimulator` instance to record events.
        The recorder subscribes to `tribal_start`, `vote_cast`, and `eliminated`.
        Call `export_xlsx(path)` after the run to write a spreadsheet.
        """
        self.sim = sim
        self.records: list[RoundRecord] = []
        # Auto-save control: if True and path provided, exporter will write
        # the spreadsheet after each elimination (end of episode).
        self.auto_save: bool = False
        self.auto_save_path: Optional[str] = None
        self._attach()

    def _attach(self) -> None:
        self.sim.events.on("tribal_start", self._on_tribal_start)
        self.sim.events.on("vote_cast", self._on_vote_cast)
        self.sim.events.on("eliminated", self._on_eliminated)

    def _on_tribal_start(self, **kwargs: Any) -> None:
        # Start a new record for this round
        rnum = _round_value(kwargs.get("round"))
        pool = _iter_objects(kwargs.get("pool"))
        rec: RoundRecord = {
            "round": rnum,
            "pool": [_name_of(p) for p in pool],
            "votes": {},  # voter_name -> target_name
            "eliminated": None,
        }
        self.records.append(rec)

    def _find_record_for_round(self, r: Optional[int]) -> Optional[RoundRecord]:
        if not self.records:
            return None
        if r is None:
            return self.records[-1]
        for rec in reversed(self.records):
            if rec["round"] == r:
                return rec
        return None

    def _on_vote_cast(self, **kwargs: Any) -> None:
        voter = kwargs.get("voter")
        target = kwargs.get("target")
        r = _round_value(kwargs.get("round"))

        found_rec = self._find_record_for_round(r)

        if found_rec is None:
            rec: RoundRecord = {
                "round": r,
                "pool": [],
                "votes": {},
                "eliminated": None,
            }
            self.records.append(rec)
        else:
            rec = found_rec

        if voter is None or target is None:
            return

        rec["votes"][_name_of(voter)] = _name_of(target)

    def _on_eliminated(self, **kwargs: Any) -> None:
        player = kwargs.get("player")
        r = _round_value(kwargs.get("round"))

        found_rec = self._find_record_for_round(r)

        if found_rec is None:
            rec: RoundRecord = {
                "round": r,
                "pool": [],
                "votes": {},
                "eliminated": None,
            }
            self.records.append(rec)
        else:
            rec = found_rec

        if player is not None:
            rec["eliminated"] = _name_of(player)

        try:
            if self.auto_save and self.auto_save_path:
                self.export_xlsx(self.auto_save_path)
        except Exception:
            pass

    # ------------------------- Exporters -------------------------
    def export_xlsx(self, path: str) -> None:
        """Export a formatted .xlsx file. Falls back to CSV matrix if openpyxl unavailable."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except Exception:
            # fallback: write a CSV-like matrix
            self._export_csv_like(path)
            return

        wb = Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet("Season")
        ws.title = "Season"

        # build player list (chronological eliminated then remaining players)
        sim = self.sim
        try:
            names = [_name_of(p) for p in _iter_objects(getattr(sim, "boot_log", None))]
            # append remaining players (finalists/winner(s))
            names += [
                name
                for name in (_name_of(p) for p in _iter_objects(getattr(sim, "players", None)))
                if name not in names
            ]
        except Exception:
            # defensive fallback: gather names from records
            seen: set[str] = set()
            names: list[str] = []
            for rec in self.records:
                for n in rec["pool"]:
                    if n not in seen:
                        seen.add(n)
                        names.append(n)

        # Header rows
        ws.cell(row=1, column=1, value="")
        ws.cell(row=2, column=1, value="Eliminated")
        ws.cell(row=3, column=1, value="Vote")

        # Episodes columns
        for j, rec in enumerate(self.records, start=2):
            col = j
            ep_label = f"{rec['round'] or (j-1)}"
            ws.cell(row=1, column=col, value=f"Episode {ep_label}")
            elim = rec["eliminated"]
            ws.cell(row=2, column=col, value=elim)
            # compute vote summary
            counts: Counter[str] = Counter(rec["votes"].values())
            if counts:
                parts = [
                    f"{target}"
                    for _, target in sorted(
                        ((count, target) for target, count in counts.items()),
                        reverse=True,
                    )
                ]
                ws.cell(row=3, column=col, value="-".join(parts))
            else:
                ws.cell(row=3, column=col, value="")

        # Player rows
        start_row = 4
        for i, pname in enumerate(names, start=start_row):
            ws.cell(row=i, column=1, value=pname)
            for j, rec in enumerate(self.records, start=2):
                col = j
                vote = rec["votes"].get(pname, "")
                ws.cell(row=i, column=col, value=vote)

        # Styling
        bold = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center")
        header_fill = PatternFill(
            start_color="FFDDDDDD", end_color="FFDDDDDD", fill_type="solid")
        elim_fill = PatternFill(start_color="FFBFBFBF",
                                end_color="FFBFBFBF", fill_type="solid")

        # Apply header styles
        for col in range(1, len(self.records) + 2):
            cell = ws.cell(row=1, column=col)
            cell.font = bold
            cell.alignment = center
            cell.fill = header_fill
            ws.row_dimensions[1].height = 18
            # second row (eliminated)
            cell2 = ws.cell(row=2, column=col)
            cell2.alignment = center
            cell2.fill = elim_fill
            cell2.font = bold
            # third row (vote summary)
            cell3 = ws.cell(row=3, column=col)
            cell3.alignment = center

        # Bold player names column
        for r in range(start_row, start_row + len(names)):
            ws.cell(row=r, column=1).font = bold

        # Auto column widths (simple heuristic)
        for col in range(1, len(self.records) + 2):
            max_len = 0
            for row in range(1, start_row + len(names)):
                val = ws.cell(row=row, column=col).value
                if val is None:
                    continue
                l = len(str(val))
                if l > max_len:
                    max_len = l
            ws.column_dimensions[get_column_letter(
                col)].width = min(40, max(10, max_len + 2))

        # Freeze panes (keep headers visible)
        ws.freeze_panes = ws["B4"]

        wb.save(path)

    def _export_csv_like(self, path: str) -> None:
        """Write a CSV-like file (tab-separated) as fallback when openpyxl not available."""
        import csv
        out_path = path if path.lower().endswith(".csv") else path + ".csv"
        # gather same names list
        sim = self.sim
        players = getattr(sim, "players", None)
        if not players:
            names: list[str] = []
            seen: set[str] = set()
            for rec in self.records:
                for p in rec["pool"]:
                    if p not in seen:
                        seen.add(p)
                        names.append(p)
        else:
            try:
                names = [_name_of(p) for p in _iter_objects(getattr(sim, "boot_log", None))]
                names += [
                    name
                    for name in (_name_of(p) for p in _iter_objects(players))
                    if name not in names
                ]
            except Exception:
                seen: set[str] = set()
                names = []
                for rec in self.records:
                    for n in rec["pool"]:
                        if n not in seen:
                            seen.add(n)
                            names.append(n)

        with open(out_path, "w", newline="", encoding="utf-8") as fh:
            w = csv.writer(fh)
            header = [
                ""] + [f"Episode {rec['round'] or idx}" for idx, rec in enumerate(self.records, start=1)]
            w.writerow(header)
            elim_row = ["Eliminated"] + \
                [rec["eliminated"] or "" for rec in self.records]
            vote_row = ["Vote"] + [
                "-".join(
                    target
                    for _, target in sorted(
                        (
                            (count, target)
                            for target, count in Counter(rec["votes"].values()).items()
                        ),
                        reverse=True,
                    )
                )
                if rec["votes"] else ""
                for rec in self.records
            ]
            w.writerow(elim_row)
            w.writerow(vote_row)
            for pname in names:
                row = [pname] + [rec["votes"].get(pname, "")
                                 for rec in self.records]
                w.writerow(row)


__all__ = ["SpreadsheetRecorder"]
